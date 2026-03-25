# Excel

import openpyxl as xl

path = ("./others/transactions.xlsx")
wordbook = xl.load_workbook(path)
sheet = wordbook["Sheet1"]

for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    print(cell.value)

